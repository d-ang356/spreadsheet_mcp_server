"""
Enhanced Spreadsheet MCP Server
- Adds formula support across sheets
- Sheet and file renaming
- Advanced cell/row/column formatting
- Freezing panes
- Pie and Bar chart generation

Usage: run as the MCP stdio server (keeps reading JSON-RPC lines from stdin)
"""

import asyncio
import json
import logging
import sys
from pathlib import Path
from typing import Optional, List, Tuple, Dict
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, PieChart, Reference

logging.basicConfig(level=logging.INFO, stream=sys.stderr)
logger = logging.getLogger("spreadsheet-mcp-server")


class SpreadsheetServer:
    def __init__(self, base_path: str = "spreadsheets", import_path: str = "/imports"):
        self.base_path = Path(base_path)
        self.base_path.mkdir(parents=True, exist_ok=True)
        self.import_path = Path(import_path)
        logger.info(f"Initialized with base path: {self.base_path}, import path: {self.import_path}")

    def _resolve_path(self, filename: str, check_exists: bool = False) -> Path:
        path = (self.base_path / filename).resolve()
        if not str(path).startswith(str(self.base_path.resolve())):
            raise ValueError("Path traversal not allowed")
        if check_exists and not path.exists():
            raise FileNotFoundError(f"File not found: {filename}")
        return path

    # ---------------------- file / sheet utilities ----------------------
    async def list_files(self, pattern: str = "*") -> dict:
        files = []
        for path in self.base_path.glob(pattern):
            if path.is_file() and path.suffix.lower() in [".xlsx", ".csv"]:
                files.append({
                    "name": path.name,
                    "size": path.stat().st_size,
                    "type": path.suffix.lower()
                })
        return {"success": True, "files": files, "count": len(files)}

    async def create_spreadsheet(self, filename: str, format: str = "xlsx",
                                 headers: Optional[list] = None,
                                 sheet_name: str = "Sheet1") -> dict:
        if not filename.endswith(f".{format}"):
            filename = f"{filename}.{format}"
        path = self._resolve_path(filename)

        if path.exists():
            return {"success": False, "error": f"File {filename} already exists"}

        if format == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            if headers:
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            wb.save(path)
        elif format == "csv":
            with open(path, 'w', newline='', encoding='utf-8') as f:
                if headers:
                    csv.writer(f).writerow(headers)
        else:
            raise ValueError(f"Unsupported format: {format}. Use 'xlsx' or 'csv'")

        return {"success": True, "filename": filename, "path": str(path)}

    async def rename_file(self, old_filename: str, new_filename: str) -> dict:
        old = self._resolve_path(old_filename, check_exists=True)
        new = self._resolve_path(new_filename)
        if new.exists():
            return {"success": False, "error": "Target filename already exists"}
        old.rename(new)
        return {"success": True, "old": old.name, "new": new.name}

    async def rename_sheet(self, filename: str, old_sheet: str, new_sheet: str) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        if old_sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet {old_sheet} not found"}
        ws = wb[old_sheet]
        ws.title = new_sheet
        wb.save(path)
        return {"success": True, "file": path.name, "sheet": new_sheet}

    # ---------------------- read / write / formula ----------------------
    async def read_spreadsheet(self, filename: str, sheet: Optional[str] = None,
                               max_rows: Optional[int] = None) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        ext = path.suffix.lower()

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(path, data_only=False)
            ws = wb[sheet] if sheet else wb.active
            data = []
            for i, row in enumerate(ws.iter_rows(values_only=False)):
                if max_rows and i >= max_rows:
                    break
                data.append([c.value for c in row])
            return {
                "success": True,
                "data": data,
                "sheet_name": ws.title,
                "rows": len(data),
                "columns": len(data[0]) if data else 0
            }
        elif ext == ".csv":
            with open(path, 'r', encoding='utf-8') as f:
                data = list(csv.reader(f))
                if max_rows:
                    data = data[:max_rows]
            return {
                "success": True,
                "data": data,
                "rows": len(data),
                "columns": len(data[0]) if data else 0
            }
        else:
            raise ValueError("Unsupported format")

    async def write_spreadsheet(self, filename: str, data: list,
                                sheet: Optional[str] = None,
                                append: bool = False) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        ext = path.suffix.lower()

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(path)

            # If sheet is specified, check if it exists, create if not
            if sheet:
                if sheet not in wb.sheetnames:
                    ws = wb.create_sheet(sheet)
                else:
                    ws = wb[sheet]
            else:
                ws = wb.active

            if not append:
                ws.delete_rows(1, ws.max_row)

            for row in data:
                ws.append(row)
            wb.save(path)
            return {"success": True, "rows_written": len(data), "sheet": ws.title}

        elif ext == ".csv":
            mode = 'a' if append else 'w'
            with open(path, mode, newline='', encoding='utf-8') as f:
                csv.writer(f).writerows(data)
            return {"success": True, "rows_written": len(data)}
        else:
            raise ValueError("Unsupported format")

    async def append_row(self, filename: str, row_data: list,
                        sheet: Optional[str] = None) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        ext = path.suffix.lower()

        if ext == ".xlsx":
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet] if sheet else wb.active
            ws.append(row_data)
            wb.save(path)
            return {"success": True, "row_number": ws.max_row}
        elif ext == ".csv":
            with open(path, 'a', newline='', encoding='utf-8') as f:
                csv.writer(f).writerow(row_data)
            return {"success": True}
        else:
            raise ValueError("Unsupported format")

    async def set_formula(self, filename: str, sheet: str, cell: str, formula: str) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        ws[cell] = formula
        wb.save(path)
        return {"success": True, "cell": cell, "formula": formula}

    async def get_formula(self, filename: str, sheet: str, cell: str) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb[sheet]
        val = ws[cell].value
        return {"success": True, "cell": cell, "formula_or_value": val}

    async def update_cell(self, filename: str, sheet: str, cell: str, value) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet] if sheet else wb.active
        ws[cell] = value
        wb.save(path)
        return {"success": True, "cell": cell, "value": value}

    async def delete_spreadsheet(self, filename: str) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        path.unlink()
        return {"success": True, "filename": filename}

    async def set_column_format(self, filename: str, sheet: str, column: str,
                                width: Optional[float] = None) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        if width:
            ws.column_dimensions[column].width = width
        wb.save(path)
        return {"success": True, "column": column}

    async def set_row_format(self, filename: str, sheet: str, row: int,
                             height: Optional[float] = None) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        if height:
            ws.row_dimensions[row].height = height
        wb.save(path)
        return {"success": True, "row": row}

    async def create_chart(self, filename: str, sheet: str, chart_type: str,
                           data_range: str, title: str = "Chart") -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            return {"success": False, "error": "Unsupported chart type"}

        chart.title = title
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "A1")
        wb.save(path)
        return {"success": True, "chart_type": chart_type}

    @staticmethod
    def _normalize_color(self, color: Optional[str]) -> Optional[str]:
        """Convert RGB hex color to ARGB format for openpyxl"""
        if not color:
            return None

        # Remove # if present
        color = color.lstrip('#')

        # If it's 6 characters (RGB), prepend FF for full opacity
        if len(color) == 6:
            return 'FF' + color.upper()
        # If it's already 8 characters (ARGB), use as-is
        elif len(color) == 8:
            return color.upper()
        else:
            raise ValueError(f"Invalid color format: {color}. Use #RRGGBB or AARRGGBB")

    async def format_cells(self, filename: str, sheet: str, cell_range: str,
                           bold: bool = False, italic: bool = False,
                           bg_color: Optional[str] = None,
                           font_size: Optional[int] = None) -> dict:
        """Format a range of cells like A1:B10"""
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]

        # Normalize color to ARGB
        normalized_color = self._normalize_color(bg_color)

        for row in ws[cell_range]:
            for cell in row:
                if bold or italic or font_size:
                    cell.font = Font(bold=bold, italic=italic, size=font_size)
                if normalized_color:
                    cell.fill = PatternFill(start_color=normalized_color, fill_type="solid")

        wb.save(path)
        return {"success": True, "range": cell_range}

    async def set_cell_format(self, filename: str, sheet: str, cell: str,
                              bold: bool = False, italic: bool = False,
                              bg_color: Optional[str] = None) -> dict:
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        target = ws[cell]

        if bold or italic:
            target.font = Font(bold=bold, italic=italic)

        # Normalize color to ARGB
        normalized_color = self._normalize_color(bg_color)
        if normalized_color:
            target.fill = PatternFill(start_color=normalized_color, fill_type="solid")

        wb.save(path)
        return {"success": True, "cell": cell}

    async def freeze_panes(self, filename: str, sheet: str, cell: str) -> dict:
        """
        Freeze panes at the specified cell.
        Examples:
        - 'A2' freezes the top row
        - 'B1' freezes the first column
        - 'B2' freezes both the top row and first column
        """
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)

        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet {sheet} not found"}

        ws = wb[sheet]
        ws.freeze_panes = cell
        wb.save(path)

        return {
            "success": True,
            "sheet": sheet,
            "freeze_cell": cell,
            "message": f"Frozen panes at {cell}"
        }

    async def unfreeze_panes(self, filename: str, sheet: str) -> dict:
        """Remove frozen panes from a sheet"""
        path = self._resolve_path(filename, check_exists=True)
        wb = openpyxl.load_workbook(path)

        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet {sheet} not found"}

        ws = wb[sheet]
        ws.freeze_panes = None
        wb.save(path)

        return {
            "success": True,
            "sheet": sheet,
            "message": "Unfrozen panes"
        }

    # ---------------------- JSON-RPC / MCP glue ----------------------

def send_response(response: dict):
    print(json.dumps(response), flush=True)


async def handle_initialize(request_id):
    send_response({
        "jsonrpc": "2.0",
        "id": request_id,
        "result": {
            "protocolVersion": "2024-11-05",
            "serverInfo": {
                "name": "spreadsheet-mcp-server",
                "version": "1.1.0"
            },
            "capabilities": {"tools": {}}
        }
    })


async def handle_tools_list(request_id):
    send_response({
        "jsonrpc": "2.0",
        "id": request_id,
        "result": {
            "tools": [
                {
                    "name": "list_files",
                    "description": "List all spreadsheet files in the base directory",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "pattern": {
                                "type": "string",
                                "description": "Glob pattern to filter files (default: '*')",
                                "default": "*"
                            }
                        }
                    }
                },
                {
                    "name": "create_spreadsheet",
                    "description": "Create a new spreadsheet file",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "format": {"type": "string", "description": "File format (xlsx or csv)", "default": "xlsx"},
                            "headers": {"type": "array", "items": {"type": "string"}, "description": "Optional header row"},
                            "sheet_name": {"type": "string", "description": "Name of the first sheet", "default": "Sheet1"}
                        },
                        "required": ["filename"]
                    }
                },
                {
                    "name": "read_spreadsheet",
                    "description": "Read data from a spreadsheet",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name (optional)"},
                            "max_rows": {"type": "integer", "description": "Maximum rows to read"}
                        },
                        "required": ["filename"]
                    }
                },
                {
                    "name": "write_spreadsheet",
                    "description": "Write data to a spreadsheet",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "data": {"type": "array", "description": "2D array of data"},
                            "sheet": {"type": "string", "description": "Sheet name (optional)"},
                            "append": {"type": "boolean", "description": "Append instead of overwrite", "default": False}
                        },
                        "required": ["filename", "data"]
                    }
                },
                {
                    "name": "append_row",
                    "description": "Append a single row to a spreadsheet",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "row_data": {"type": "array", "description": "Row data as array"},
                            "sheet": {"type": "string", "description": "Sheet name (optional)"}
                        },
                        "required": ["filename", "row_data"]
                    }
                },
                {
                    "name": "update_cell",
                    "description": "Update a single cell value",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "cell": {"type": "string", "description": "Cell reference (e.g., 'A1')"},
                            "value": {"description": "Value to set"}
                        },
                        "required": ["filename", "sheet", "cell", "value"]
                    }
                },
                {
                    "name": "delete_spreadsheet",
                    "description": "Delete a spreadsheet file",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"}
                        },
                        "required": ["filename"]
                    }
                },
                {
                    "name": "format_cells",
                    "description": "Format a range of cells",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "cell_range": {"type": "string", "description": "Cell range (e.g., 'A1:B10')"},
                            "bold": {"type": "boolean", "default": False},
                            "italic": {"type": "boolean", "default": False},
                            "bg_color": {"type": "string", "description": "Background color hex (e.g., '#00FF00' or 'FF00FF00')"},
                            "font_size": {"type": "integer", "description": "Font size"}
                        },
                        "required": ["filename", "sheet", "cell_range"]
                    }
                },
                {
                    "name": "set_formula",
                    "description": "Set a formula in a cell",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "cell": {"type": "string", "description": "Cell reference"},
                            "formula": {"type": "string", "description": "Excel formula"}
                        },
                        "required": ["filename", "sheet", "cell", "formula"]
                    }
                },
                {
                    "name": "get_formula",
                    "description": "Get formula from a cell",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "cell": {"type": "string", "description": "Cell reference"}
                        },
                        "required": ["filename", "sheet", "cell"]
                    }
                },
                {
                    "name": "rename_sheet",
                    "description": "Rename a sheet",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "old_sheet": {"type": "string", "description": "Current sheet name"},
                            "new_sheet": {"type": "string", "description": "New sheet name"}
                        },
                        "required": ["filename", "old_sheet", "new_sheet"]
                    }
                },
                {
                    "name": "rename_file",
                    "description": "Rename a file",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "old_filename": {"type": "string", "description": "Current filename"},
                            "new_filename": {"type": "string", "description": "New filename"}
                        },
                        "required": ["old_filename", "new_filename"]
                    }
                },
                {
                    "name": "set_cell_format",
                    "description": "Format a single cell",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "cell": {"type": "string", "description": "Cell reference"},
                            "bold": {"type": "boolean", "default": False},
                            "italic": {"type": "boolean", "default": False},
                            "bg_color": {"type": "string", "description": "Background color hex (e.g., '#00FF00' or 'FF00FF00')"}
                        },
                        "required": ["filename", "sheet", "cell"]
                    }
                },
                {
                    "name": "set_column_format",
                    "description": "Format a column",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "column": {"type": "string", "description": "Column letter (e.g., 'A')"},
                            "width": {"type": "number", "description": "Column width"}
                        },
                        "required": ["filename", "sheet", "column"]
                    }
                },
                {
                    "name": "set_row_format",
                    "description": "Format a row",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "row": {"type": "integer", "description": "Row number"},
                            "height": {"type": "number", "description": "Row height"}
                        },
                        "required": ["filename", "sheet", "row"]
                    }
                },
                {
                    "name": "create_chart",
                    "description": "Create a chart in the spreadsheet",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "chart_type": {"type": "string", "description": "Chart type (bar or pie)"},
                            "data_range": {"type": "string", "description": "Data range (e.g., 'A1:B10')"},
                            "title": {"type": "string", "description": "Chart title", "default": "Chart"}
                        },
                        "required": ["filename", "sheet", "chart_type", "data_range"]
                    }
                },
                {
                    "name": "freeze_panes",
                    "description": "Freeze rows and/or columns at a specific cell position. Use 'A2' to freeze top row, 'B1' to freeze first column, 'B2' to freeze both.",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"},
                            "cell": {"type": "string", "description": "Cell reference where to freeze (e.g., 'A2' for top row, 'B1' for first column, 'B2' for both)"}
                        },
                        "required": ["filename", "sheet", "cell"]
                    }
                },
                {
                    "name": "unfreeze_panes",
                    "description": "Remove frozen panes from a sheet",
                    "inputSchema": {
                        "type": "object",
                        "properties": {
                            "filename": {"type": "string", "description": "Name of the file"},
                            "sheet": {"type": "string", "description": "Sheet name"}
                        },
                        "required": ["filename", "sheet"]
                    }
                }
            ]
        }
    })


async def handle_tool_call(request_id, server: SpreadsheetServer, tool_name: str, arguments: dict):
    try:
        method = getattr(server, tool_name, None)
        if not method:
            raise ValueError(f"Unknown tool: {tool_name}")

        result = await method(**(arguments or {}))
        send_response({
            "jsonrpc": "2.0",
            "id": request_id,
            "result": {"content": [{"type": "text", "text": json.dumps(result)}]}
        })

    except Exception as e:
        logger.error(f"Error in tool {tool_name}: {e}", exc_info=True)
        send_response({
            "jsonrpc": "2.0",
            "id": request_id,
            "error": {"code": -32000, "message": str(e)}
        })


async def main():
    server = SpreadsheetServer()
    logger.info("Spreadsheet MCP Server starting (enhanced)...")

    loop = asyncio.get_event_loop()

    while True:
        line = await loop.run_in_executor(None, sys.stdin.readline)

        if not line:  # EOF = shutdown
            logger.info("STDIN closed — shutting down.")
            break

        line = line.strip()
        if not line:
            continue

        request = json.loads(line)
        method = request.get("method")
        request_id = request.get("id")

        if method == "initialize":
            await handle_initialize(request_id)
        elif method == "tools/list":
            await handle_tools_list(request_id)
        elif method == "tools/call":
            tool_name = request["params"]["name"]
            arguments = request["params"].get("arguments", {})
            await handle_tool_call(request_id, server, tool_name, arguments)


if __name__ == "__main__":
    asyncio.run(main())
